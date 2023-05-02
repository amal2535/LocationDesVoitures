import re
from asyncore import write
from datetime import datetime
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv
#creation Fichier CSV et txt
def recherche_marque(ch,n):
    wb = load_workbook('Client.xlsx')
    ws = wb.active
    row=0
    char=''
    val=''
    char2=int(ch)
    for row in range(2,n):
            char=get_column_letter(1)
            val=ws[char+str(row)].value
            str(val)
            print(type(val))
            print(type(ch))
                  
            if(char2==val):
                test = row
                break
    return(test) 

def recherche_ADR(ch,n):
    wb = load_workbook('Client.xlsx')
    ws = wb.active
    row=0
    char=''
    val=''
    test = -1
    for row in range(1,n+1):
            char=get_column_letter(1)
            val=ws[char+str(row)].value
            val2=str(val)     
            if(ch==val2):
                test = row
                break
    return(test) 

def recherche_cin(ch,n):
    wb = load_workbook('Client.xlsx')
    ws = wb.active
    L=[]
    for i in range(1,8):
        char=get_column_letter(i)
        val=ws[char+str(n)].value
        L.append(val)  
    return L

def supp_c(col):
    wb = load_workbook('Client.xlsx')
    ws = wb.active
    ws.delete_rows(col,1)
    wb.save("Client.xlsx")


def remplir_csv(L):
    wb = load_workbook('Client.xlsx')
    ws = wb.active
    ws.append(L)
    wb.save("Client.xlsx")

def ecrire (ch) :
    f= open("CIN.txt",'a')
    f.write(ch)
    f.write('\n')
    f.close()

def lecture (ch) :
    list=[]
    test=2
    file=open("CIN.txt",'r')
    f=file.readlines()
    for i in f:
        list.append(i[:-1])
    if(ch in list):
        test=1
    else:
        test=0
    file.close()
    return (test)

def modif_Adresse(ch,n):
    wb = load_workbook('Client.xlsx')
    ws = wb.active
    ws.cell(row=n,column=6,value=ch)
    ch1=wb.save("Client.xlsx")


#---------------------------------
def adresse1(ch):
    if(ch[0].isupper() and ch.isalpha()):
        return 1
    else:
        return 0
    
def saisie_age():
    while(True):
        age=int(input("Donner votre age: "))
        if(age>=18 and age<80):
            break
    return age

def saisie_cin():
    while True:
        cin=input("Donner votre cin: ")
        if (cin.isdigit()) and (cin[0]=='1' or cin[0]=='0') and len(cin)==8:
            break
    return cin


def np(ch):
    if(len(ch)<20 and len(ch)>1 and ch[0].isupper() and ch.isalpha()):
        return(1)
    else:
        return(0)

def ellisa(ch):
    if (ch[0]=='4'):
        return 1
    else:
        return 0

def ooredoo(ch):
    if(ch[0]=='2'):
        return 1
    else:
        return 0

def orange(ch):
    if(ch[0]=='5'):
        return 1
    else:
        return 0

def Telec(ch):
    if(ch[0]=='9'):
        return 1
    else:
        return 0
    
def adr_gm(ch):
    regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
    if(re.search(regex,ch)):
        return 1
    else:
        return 0
    
from asyncore import write
from datetime import datetime
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv
def matricule(mat):
    tm1 = 0
    tm2 = 0
    tm3 = 0
    mat1=mat[0:3]
    mat2=mat[4:6]
    mat3=mat[7:]
    if(mat1>'100'):
        tm1=1
    if(mat2=='TN'):
        tm2=1
    if(mat3>'1000'):
        tm3=1
    return(tm1==tm2==tm3==1)

def date(ch):
    format = "%d-%m-%Y"
    res = True
    try:
        res = bool(datetime.strptime(ch, format))
    except ValueError:
        res = False
    return(res)

def remplir_csv(L):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    ws.append(L)
    wb.save("test.xlsx")

def recherche_mat(ch,n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    L=[]
    for i in range(2,n+1):
        char=get_column_letter(i)
        val=ws[char+str(n)].value
        if(ch==val):
            L.append(i)  
    return L


def supp_v(col):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    ws.delete_rows(col,1)
    wb.save("test.xlsx")


def ecrire (ch) :
    f= open("newfile.txt",'a')
    f.write(ch)
    f.write('\n')
    f.close()

def lecture (ch) :
    list=[]
    test=2
    file=open("newfile.txt",'r')
    f=file.readlines()
    for i in f:
        list.append(i[:-1])
    if(ch in list):
        test=1
    else:
        test=0
    file.close()
    return (test)

def recherche_marque(ch,n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    row=0
    char=''
    val=''
    test = -1
    for row in range(2,n):
            char=get_column_letter(2)
            val=ws[char+str(row)].value
            if(ch==val):
                test = row
                break
    return(test) 


def fromdatetoage (x) :
    today = datetime.now().date()
    #x=datetime(1989,3,15)
    return today.year - x.year - ((today.month,today.day)<(x.month,x.day))

#print(fromdatetoage())

def supp_date(n):
    wb = load_workbook('test2.xlsx')
    ws = wb.active
    test = 0
    age=0
    x=n
    for row in range(1,x):
            char=get_column_letter(5)
            val=ws[char+str(row)].value
            age=fromdatetoage(val)
            if(age>=10):
                ws.delete_rows(row,1)
                x-=1
    wb.save("test2.xlsx")

def modif_prix(ch,n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    ws.cell(row=n,column=6,value=ch)
    ch1=wb.save("test.xlsx")

def modif_couleur(ch,n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    ws.cell(row=n,column=3,value=ch)
    ch1=wb.save("test.xlsx")
    

#lES FONCTIONS DE RECHERCHE
def recherche_matricule(ch,n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    L=[]
    for i in range(1,7):
        char=get_column_letter(i)
        val=ws[char+str(i)].value
        L.append(val)  
    return L

def recherche_e(ch,n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    L=[]
    for i in range(1,n+1):
        char=get_column_letter(4)
        val=ws[char+str(i)].value
        if(ch==val):
            L.append(i)  
    return L

def recherche_etat(n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    char=''
    L=[]
    for i in range(1,7):
        char=get_column_letter(i)
        val=ws[char+str(n)].value
        L.append(val)  
    return L
def recherche_marque1(ch,n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    L=[]
    for i in range(1,n+1):
        char=get_column_letter(2)
        val=ws[char+str(i)].value
        if(ch==val):
            L.append(i)  
    return L


def recherche_mar1(n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    char=''
    L=[]
    for i in range(1,7):
        char=get_column_letter(i)
        val=ws[char+str(n)].value
        L.append(val)  
    return L

def recherche_coul1(ch,n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    L=[]
    for i in range(1,n+1):
        char=get_column_letter(3)
        val=ws[char+str(i)].value
        if(ch==val):
            L.append(i)  
    return L

def recherche_couleur1(n):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    char=''
    L=[]
    for i in range(1,7):
        char=get_column_letter(i)
        val=ws[char+str(n)].value
        L.append(val)  
    return L